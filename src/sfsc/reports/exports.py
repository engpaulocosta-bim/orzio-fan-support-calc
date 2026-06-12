"""Exportação para Excel e CSV."""

from __future__ import annotations

import csv
import io
from pathlib import Path

from .. import __version__
from ..assessment import assess_result
from ..config import get_assumptions
from ..models import ReportContext

_DISCLAIMER = (
    "ENGINEERING ESTIMATE ONLY — requer revisão por engenheiro estrutural "
    "qualificado antes de uso em projecto ou construção."
)


def generate_excel(ctx: ReportContext, output_path: str | Path | None = None) -> bytes:
    """Gera workbook Excel com folhas: Resumo, Secção, Mesa, Ancoragens, Combinações."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl") from exc

    wb = openpyxl.Workbook()
    inp = ctx.fan_support_input
    res = ctx.fan_support_result
    assessment = assess_result(res) if res else None

    BLUE_HEX = "1447E6"
    LIGHT_HEX = "EFF4FE"
    GREY_HEX = "64748B"

    def _header_style(cell, bg=BLUE_HEX):
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _set_col_widths(ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _kv_row(ws, row, key, value, row_n):
        ws.cell(row=row_n, column=1, value=key).font = Font(bold=True, size=9)
        ws.cell(row=row_n, column=1).fill = PatternFill("solid", fgColor=LIGHT_HEX)
        ws.cell(row=row_n, column=2, value=value).font = Font(size=9)
        ws.cell(row=row_n, column=2).alignment = Alignment(wrap_text=True)

    # ── Folha 1: Resumo ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws.cell(1, 1, "SFSC — Steel Fan Support Calc").font = Font(bold=True, size=14, color=BLUE_HEX)
    ws.cell(2, 1, f"Projecto: {ctx.project_name}  |  Tag: {ctx.support_tag}  |  {ctx.date}")
    ws.cell(2, 1).font = Font(size=9, color=GREY_HEX)
    n = 4
    kv_data = []
    if inp:
        kv_data += [
            ("Tipo de suporte", inp.support_type.value),
            ("País / norma", f"{inp.country.value}"),
            ("Aço", inp.steel_grade.value),
            ("Peso total operação [kg]", inp.total_operating_weight_kg),
            ("Carga de projecto [kN]", res.design_load_kN if res else "—"),
            ("Factor sísmico ag/g", res.seismic_factor_g if res else "—"),
        ]
    if res:
        kv_data += [
            ("Diagnóstico", assessment.headline if assessment else "—"),
            ("Verificação governante", assessment.governing_item if assessment else "—"),
            (
                "Utilização governante [%]",
                round(assessment.utilization_percent, 1)
                if assessment and assessment.utilization_percent is not None
                else "—",
            ),
            (
                "Conservadorismo informativo [%]",
                round(assessment.conservatism_percent, 1)
                if assessment and assessment.conservatism_percent is not None
                else "—",
            ),
            (
                "Perfil seleccionado",
                res.recommended_section.designation if res.recommended_section else "—",
            ),
            (
                "Utilização máx. secção",
                res.section_verification.utilization_ratio if res.section_verification else "—",
            ),
            ("Estado global", res.status.value),
            ("Classificação", res.classification_level.value),
        ]
    for k, v in kv_data:
        _kv_row(ws, None, k, v, n)
        n += 1
    _set_col_widths(ws, [35, 30])

    # ── Folha 2: Combinações ──────────────────────────────────────────────────
    # Dois níveis distintos (auditoria C-02): acções totais vs esforços no elemento.
    ws2 = wb.create_sheet("Combinações")
    headers_c = ["Nível", "Combinação", "V_z (kN)", "V_y (kN)", "M_y (kNm)", "N (kN)", "Governante"]
    for j, h in enumerate(headers_c, 1):
        _header_style(ws2.cell(1, j, h))
    row_n = 2
    if res:
        for level, combos in (
            ("Acções totais", res.all_combinations),
            ("Esforços no elemento", res.member_forces),
        ):
            for c in combos:
                ws2.cell(row_n, 1, level)
                ws2.cell(row_n, 2, c.name)
                ws2.cell(row_n, 3, round(c.V_z_kN, 3))
                ws2.cell(row_n, 4, round(c.V_y_kN, 3))
                ws2.cell(row_n, 5, round(c.M_y_kNm, 3))
                ws2.cell(row_n, 6, round(c.N_kN, 3))
                ws2.cell(row_n, 7, "Sim" if c.governing else "")
                row_n += 1
    _set_col_widths(ws2, [22, 28, 16, 16, 16, 16, 14])

    # ── Folha 3: Secção ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Secção")
    if res and res.recommended_section:
        sec = res.recommended_section
        n3 = 1
        for k, v in [
            ("Designação", sec.designation),
            ("Família", sec.family.value),
            ("h [mm]", sec.h_mm),
            ("b [mm]", sec.b_mm),
            ("tw [mm]", sec.tw_mm),
            ("tf [mm]", sec.tf_mm),
            ("A [cm²]", sec.A_cm2),
            ("I_y [cm⁴]", sec.I_y_cm4),
            ("W_el,y [cm³]", sec.W_el_y_cm3),
            ("Peso [kg/m]", sec.weight_kgm),
        ]:
            _kv_row(ws3, None, k, v, n3)
            n3 += 1

        if res.section_verification:
            sv = res.section_verification
            n3 += 1
            ws3.cell(n3, 1, "Verificações").font = Font(bold=True, color=BLUE_HEX)
            n3 += 1
            headers_v = ["Check", "η (utilização)", "Estado"]
            for j, h in enumerate(headers_v, 1):
                _header_style(ws3.cell(n3, j, h))
            n3 += 1
            for check, eta in sv.utilization_by_check.items():
                ws3.cell(n3, 1, check)
                ws3.cell(n3, 2, round(eta, 4))
                ws3.cell(n3, 3, "OK" if eta <= 1.0 else "FALHA")
                if eta > 1.0:
                    ws3.cell(n3, 3).fill = PatternFill("solid", fgColor="FECACA")
                n3 += 1
    _set_col_widths(ws3, [28, 20, 16])

    # ── Folha 4: Mesa ─────────────────────────────────────────────────────────
    if res and res.base_plate:
        ws4 = wb.create_sheet("Mesa (Base Plate)")
        bp = res.base_plate
        n4 = 1
        for k, v in [
            ("L [mm]", bp.length_mm),
            ("B [mm]", bp.width_mm),
            ("Espessura [mm]", bp.thickness_mm),
            ("Aço", bp.steel_grade.value),
            ("σ_bearing [MPa]", bp.bearing_stress_mpa),
            ("η_bearing", bp.utilization_bearing),
            ("η_bending", bp.utilization_bending),
            ("Parafusos ventilador", f"M{bp.bolt_diameter_mm:.0f} × {bp.n_bolts_fan}"),
            ("η_bolt_fan", bp.bolt_utilization_fan),
            ("Parafusos estrutura", f"M20 × {bp.n_bolts_structure}"),
            ("η_bolt_str", bp.bolt_utilization_structure),
            ("Furo ancoragem [mm]", bp.hole_diameter_mm),
            ("Espaçamento X [mm]", bp.anchor_spacing_x_mm),
            ("Espaçamento Y [mm]", bp.anchor_spacing_y_mm),
            ("Espaçamento mín. [mm]", bp.min_spacing_mm),
            ("Bordo X [mm]", bp.edge_distance_x_mm),
            ("Bordo Y [mm]", bp.edge_distance_y_mm),
            ("Bordo mín. [mm]", bp.min_edge_distance_mm),
            ("η cone betão", bp.utilization_concrete_cone),
            ("η pull-out", bp.utilization_pullout),
            ("η pry-out", bp.utilization_pryout),
            ("Garganta soldadura [mm]", bp.weld_throat_mm),
            ("η_weld", bp.weld_utilization),
            ("Estado", bp.status.value),
        ]:
            _kv_row(ws4, None, k, v, n4)
            n4 += 1
        _set_col_widths(ws4, [35, 25])

    # ── Folha 5: Ancoragens / varões ──────────────────────────────────────────
    if res and res.anchor:
        anc = res.anchor
        is_rod = anc.anchor_type == "rod"
        ws5 = wb.create_sheet("Varões Suspensão" if is_rod else "Ancoragens")
        rows5 = [
            (
                "Tipo de verificação",
                "Varão roscado de suspensão (sem betão)"
                if is_rod
                else "Ancoragem embebida em betão",
            ),
            ("Nº", anc.n_anchors),
            ("Diâmetro [mm]", anc.anchor_diameter_mm),
        ]
        if not is_rod:
            rows5.append(("Profundidade hef [mm]", anc.embedment_depth_mm))
        rows5 += [
            ("N_Rd total [kN]", anc.tensile_capacity_kN),
            ("V_Rd total [kN]", anc.shear_capacity_kN),
            ("η_tracção", anc.utilization_tension),
            ("η_corte", anc.utilization_shear),
            ("η_interacção", anc.utilization_combined),
            ("Estado", anc.status.value),
            ("Cláusula", anc.code_clause),
        ]
        n5 = 1
        for k, v in rows5:
            _kv_row(ws5, None, k, v, n5)
            n5 += 1
        _set_col_widths(ws5, [35, 25])

    # ── Folha 6: Ligações metálicas ───────────────────────────────────────────
    if res and res.metal_connection:
        ws6 = wb.create_sheet("Ligações Metálicas")
        mc = res.metal_connection
        n6 = 1
        rows6: list[tuple[str, object]] = [
            ("Tipo", mc.connection_type),
            ("Chapa ligação [mm]", mc.plate_thickness_mm),
            ("Parafusos", f"M{mc.bolt_diameter_mm:.0f} × {mc.n_bolts} ({mc.bolt_grade})"),
            ("Furo [mm]", mc.bolt_hole_diameter_mm),
            ("Espaçamento fornecido [mm]", mc.provided_spacing_mm),
            ("Espaçamento mínimo [mm]", mc.min_spacing_mm),
            ("Bordo fornecido [mm]", mc.provided_edge_distance_mm),
            ("Bordo mínimo [mm]", mc.min_edge_distance_mm),
            ("N_Rd corte parafusos [kN]", mc.bolt_shear_capacity_kN),
            ("N_Rd tração parafusos [kN]", mc.bolt_tension_capacity_kN),
            ("η corte parafusos", mc.utilization_bolt_shear),
            ("η tração parafusos", mc.utilization_bolt_tension),
            ("Garganta soldadura [mm]", mc.weld_throat_mm),
            ("Comprimento soldadura [mm]", mc.weld_length_mm),
            ("η soldadura", mc.weld_utilization),
            (
                "Stiffener",
                f"{mc.stiffener_thickness_mm:.1f} mm" if mc.stiffener_required else "Não",
            ),
            ("Cantoneira", mc.cleat_angle or "—"),
            ("Diagonal", mc.diagonal_member or "—"),
            ("η global", mc.utilization_ratio),
            ("Estado", mc.status.value),
            ("Cláusula", mc.code_clause),
        ]
        for k6, v6 in rows6:
            _kv_row(ws6, None, k6, v6, n6)
            n6 += 1
        _set_col_widths(ws6, [38, 35])

    # Folha de quantidades preliminares (Fase 5 / orcamentacao)
    if res and res.quantities:
        wsq = wb.create_sheet("Quantidades")
        q = res.quantities
        nq = 1
        for kq, vq in [
            ("Nota", q.estimate_note),
            ("Comprimento perfis [m]", q.structural_steel_length_m),
            ("Massa perfis [kg]", q.structural_steel_mass_kg),
            ("Massa chapas [kg]", q.plate_mass_kg),
            ("Massa total aco [kg]", q.total_steel_mass_kg),
            ("Ancoragens/varoes [n]", q.anchor_count),
            ("Diametro ancoragem/varao [mm]", q.anchor_diameter_mm),
            ("Embebimento/comprimento [mm]", q.anchor_embedment_or_length_mm),
            ("Tipo ancoragem", q.anchor_type),
            ("Comprimento solda [mm]", q.weld_length_mm),
            ("Garganta solda [mm]", q.weld_throat_mm),
        ]:
            _kv_row(wsq, None, kq, vq, nq)
            nq += 1

        nq += 1
        for j, h in enumerate(["Item", "Quantidade", "Comprimento", "Massa [kg]", "Base"], 1):
            _header_style(wsq.cell(nq, j, h))
        nq += 1
        for item in q.line_items:
            wsq.cell(nq, 1, str(item.get("item", "")))
            wsq.cell(nq, 2, item.get("quantity", ""))
            wsq.cell(
                nq,
                3,
                item.get("length_m", item.get("length_mm", item.get("embedment_or_length_mm", ""))),
            )
            wsq.cell(nq, 4, item.get("mass_kg", ""))
            wsq.cell(nq, 5, item.get("section", item.get("type", "")))
            nq += 1
        _set_col_widths(wsq, [36, 16, 18, 16, 24])

    # ── Folha 7: Avisos e Pressupostos (auditoria H-04) ──────────────────────
    ws7 = wb.create_sheet("Avisos e Pressupostos")
    n7 = 1
    ws7.cell(n7, 1, "Avisos").font = Font(bold=True, size=12, color=BLUE_HEX)
    n7 += 1
    for j, h in enumerate(["Código", "Severidade", "Mensagem", "Módulo"], 1):
        _header_style(ws7.cell(n7, j, h))
    n7 += 1
    for w in ctx.warnings:
        ws7.cell(n7, 1, w.code)
        sev_cell = ws7.cell(n7, 2, w.severity)
        if w.severity == "CRITICAL":
            sev_cell.font = Font(bold=True, color="B91C1C")
        ws7.cell(n7, 3, w.message).alignment = Alignment(wrap_text=True)
        ws7.cell(n7, 4, w.module)
        n7 += 1

    n7 += 1
    ws7.cell(n7, 1, "Pressupostos declarados").font = Font(bold=True, size=12, color=BLUE_HEX)
    n7 += 1
    assumption_index = {
        a.get("id"): a.get("description", "") for a in get_assumptions().get("assumptions", [])
    }
    for j, h in enumerate(["ID", "Descrição"], 1):
        _header_style(ws7.cell(n7, j, h))
    n7 += 1
    for a_id in ctx.assumptions_declared:
        ws7.cell(n7, 1, a_id)
        ws7.cell(n7, 2, assumption_index.get(a_id, "")).alignment = Alignment(wrap_text=True)
        n7 += 1

    n7 += 1
    ws7.cell(n7, 1, "Limitações").font = Font(bold=True, size=12, color=BLUE_HEX)
    n7 += 1
    for lim in ctx.limitations:
        ws7.cell(n7, 1, lim).alignment = Alignment(wrap_text=True)
        n7 += 1
    _set_col_widths(ws7, [16, 14, 90, 18])

    # ── Folha 8: Info (versão + proveniência dos datasets — auditoria H-07) ──
    ws8 = wb.create_sheet("Info")
    provenance = ctx.dataset_provenance or {}
    n8 = 1
    info_rows: list[tuple[str, object]] = [
        ("Software", f"SFSC v{__version__}"),
        ("Projecto", ctx.project_name),
        ("Tag", ctx.support_tag),
        ("Data", ctx.date),
        ("Revisão", ctx.revision),
        ("Engenheiro responsável", ctx.prepared_by),
        ("Hash combinado dos datasets", str(provenance.get("datasets_combined_sha256", ""))[:12]),
        ("Disclaimer", _DISCLAIMER),
    ]
    for k8, v8 in info_rows:
        _kv_row(ws8, None, k8, v8, n8)
        n8 += 1
    n8 += 1
    for j, h in enumerate(["Dataset", "SHA-256", "Modificado"], 1):
        _header_style(ws8.cell(n8, j, h))
    n8 += 1
    for name, meta in provenance.get("datasets", {}).items():
        ws8.cell(n8, 1, name)
        ws8.cell(n8, 2, str(meta.get("sha256", ""))[:12])
        ws8.cell(n8, 3, str(meta.get("modified", "—")))
        n8 += 1
    _set_col_widths(ws8, [38, 35, 16])

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    if output_path:
        Path(output_path).write_bytes(xlsx_bytes)
    return xlsx_bytes


def generate_csv(ctx: ReportContext, output_path: str | Path | None = None) -> str:
    """Gera CSV resumo de uma linha (útil para batch)."""
    inp = ctx.fan_support_input
    res = ctx.fan_support_result
    assessment = assess_result(res) if res else None

    fields = {
        "support_tag": ctx.support_tag,
        "project": ctx.project_name,
        "diagnosis": assessment.headline if assessment else "",
        "governing_item": assessment.governing_item if assessment else "",
        "governing_utilization_pct": assessment.utilization_percent if assessment else "",
        "conservatism_pct": assessment.conservatism_percent if assessment else "",
        "limit_usage_pct": assessment.limit_percent if assessment else "",
        "exceedance_pct": assessment.exceedance_percent if assessment else "",
        "support_type": inp.support_type.value if inp else "",
        "country": inp.country.value if inp else "",
        "total_weight_kg": inp.total_operating_weight_kg if inp else "",
        "design_load_kN": res.design_load_kN if res else "",
        "seismic_factor_g": res.seismic_factor_g if res else "",
        "section": res.recommended_section.designation if (res and res.recommended_section) else "",
        "section_utilization": res.section_verification.utilization_ratio
        if (res and res.section_verification)
        else "",
        "base_plate_t_mm": res.base_plate.thickness_mm if (res and res.base_plate) else "",
        "base_plate_hole_d_mm": res.base_plate.hole_diameter_mm if (res and res.base_plate) else "",
        "base_plate_concrete_cone_eta": res.base_plate.utilization_concrete_cone
        if (res and res.base_plate)
        else "",
        "base_plate_pullout_eta": res.base_plate.utilization_pullout
        if (res and res.base_plate)
        else "",
        "base_plate_pryout_eta": res.base_plate.utilization_pryout
        if (res and res.base_plate)
        else "",
        "anchor_d_mm": res.anchor.anchor_diameter_mm if (res and res.anchor) else "",
        "n_anchors": res.anchor.n_anchors if (res and res.anchor) else "",
        "anchor_type": res.anchor.anchor_type if (res and res.anchor) else "",
        "metal_connection": res.metal_connection.connection_type
        if (res and res.metal_connection)
        else "",
        "metal_connection_eta": res.metal_connection.utilization_ratio
        if (res and res.metal_connection)
        else "",
        "metal_connection_bolts": (
            f"M{res.metal_connection.bolt_diameter_mm:.0f}x{res.metal_connection.n_bolts}"
            if (res and res.metal_connection)
            else ""
        ),
        "quantity_total_steel_kg": res.quantities.total_steel_mass_kg
        if (res and res.quantities)
        else "",
        "quantity_anchor_count": res.quantities.anchor_count if (res and res.quantities) else "",
        "quantity_weld_length_mm": res.quantities.weld_length_mm
        if (res and res.quantities)
        else "",
        "status": res.status.value if res else "",
        "classification": res.classification_level.value if res else "",
        "warnings_count": len(ctx.warnings),
        "critical_warnings": sum(1 for w in ctx.warnings if w.severity == "CRITICAL"),
        "software_version": __version__,
        "datasets_sha256": str((ctx.dataset_provenance or {}).get("datasets_combined_sha256", ""))[
            :12
        ],
        "prepared_by": ctx.prepared_by,
        "disclaimer": _DISCLAIMER,
    }
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(fields.keys()))
    writer.writeheader()
    writer.writerow(fields)
    csv_str = buf.getvalue()
    if output_path:
        Path(output_path).write_text(csv_str, encoding="utf-8")
    return csv_str
