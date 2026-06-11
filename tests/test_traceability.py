"""Rastreabilidade e relatórios da Fase 3: provenance, versão, dedupe de
citações, memória de fórmulas, modo VERIFY e escolha manual de perfil."""

import base64
import csv
import io
import re
import zlib

import openpyxl

from sfsc import __version__
from sfsc.config import get_dataset_provenance
from sfsc.engines.selector import context_for_section_choice, run_full_calculation
from sfsc.enums import Country, FanType, OperationMode, SectionFamily, SupportType
from sfsc.models import FanSupportInput, FanUnit
from sfsc.reports.exports import generate_csv, generate_excel
from sfsc.reports.memorial_pdf import generate_pdf


def _pdf_text(pdf_bytes: bytes) -> bytes:
    text = b""
    for m in re.finditer(rb"(?<!end)stream\r?\n(.*?)endstream", pdf_bytes, re.DOTALL):
        data = m.group(1).strip()
        try:
            text += zlib.decompress(base64.a85decode(data, adobe=True))
        except Exception:
            try:
                text += zlib.decompress(data)
            except Exception:
                pass
    return text


def _inp(**kwargs):
    defaults = dict(
        project_name="Trace",
        support_tag="FSU-T",
        prepared_by="Eng. Responsável",
        fan_units=[
            FanUnit(
                fan_type=FanType.CENTRIFUGAL,
                weight_kg=120.0,
                operating_weight_kg=130.0,
                footprint_length_mm=800.0,
                footprint_width_mm=600.0,
            )
        ],
        support_type=SupportType.HANGER,
        country=Country.PORTUGAL,
        seismic_zone="1.3",
        installation_height_mm=500.0,
        span_mm=1200.0,
    )
    defaults.update(kwargs)
    return FanSupportInput(**defaults)


# ── Proveniência ────────────────────────────────────────────────────────────


def test_provenance_structure_and_hashes():
    prov = get_dataset_provenance()
    assert prov["software_version"] == __version__
    assert len(prov["datasets_combined_sha256"]) == 64
    # Cada dataset shipped tem hash de 64 hex (não MISSING).
    for name, meta in prov["datasets"].items():
        assert len(meta["sha256"]) == 64, name


def test_report_context_carries_provenance():
    ctx = run_full_calculation(_inp())
    assert ctx.dataset_provenance.get("datasets_combined_sha256")


def test_provenance_in_all_outputs():
    ctx = run_full_calculation(_inp())
    short = ctx.dataset_provenance["datasets_combined_sha256"][:12]

    csv_row = next(csv.DictReader(io.StringIO(generate_csv(ctx))))
    assert csv_row["datasets_sha256"] == short
    assert csv_row["software_version"] == __version__
    assert csv_row["warnings_count"].isdigit()
    assert "ENGINEERING ESTIMATE" in csv_row["disclaimer"]

    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(ctx)))
    assert "Info" in wb.sheetnames
    assert "Avisos e Pressupostos" in wb.sheetnames
    info_vals = " ".join(
        str(c.value) for row in wb["Info"].iter_rows() for c in row if c.value is not None
    )
    assert short in info_vals
    assert __version__ in info_vals

    pdf = _pdf_text(generate_pdf(ctx))
    assert b"13. RASTREABILIDADE" in pdf
    assert short.encode() in pdf


# ── Dedupe de citações por (norma, cláusula) — H-06 ──────────────────────────


def test_citations_keep_distinct_clauses_same_standard():
    # Pedestal com base plate gera EN1993-1-8 em duas cláusulas distintas
    # (6.2.5 base plate e 3+4+6 ligações) — ambas devem sobreviver.
    from sfsc.enums import FanConnectionType

    ctx = run_full_calculation(
        _inp(
            support_type=SupportType.PEDESTAL,
            include_base_plate=True,
            fan_connection_type=FanConnectionType.DIRECT_FLANGE,
        )
    )
    en1998 = [c for c in ctx.citations if c.standard_id == "EN1993-1-8"]
    clauses = {c.clause for c in en1998}
    assert len(clauses) >= 2, clauses


# ── Memória de fórmulas ──────────────────────────────────────────────────────


def test_calculation_details_populated():
    res = run_full_calculation(_inp()).fan_support_result
    d = res.section_verification.calculation_details
    # LTB governa neste caso → grandezas do LTB presentes.
    assert "Mb_Rd_kNm" in d
    assert "lambda_LT" in d
    assert d["fy_MPa"] == 355


def test_formula_memory_in_pdf():
    pdf = _pdf_text(generate_pdf(run_full_calculation(_inp())))
    assert b"Mem" in pdf  # "Memória de fórmulas"
    assert b"Mb,Rd" in pdf or b"Vpl,Rd" in pdf


# ── Modo VERIFY ──────────────────────────────────────────────────────────────


def test_verify_mode_runs_against_given_section():
    ctx = run_full_calculation(
        _inp(
            operation_mode=OperationMode.VERIFY,
            received_section_family=SectionFamily.HEB,
            received_section_tag="HEB160",
        )
    )
    res = ctx.fan_support_result
    assert res.recommended_section.designation == "HEB160"
    assert len(res.section_options) == 1


# ── Escolha manual de perfil regista nota (M-04) ─────────────────────────────


def test_manual_section_choice_records_note():
    ctx = run_full_calculation(_inp())
    options = ctx.fan_support_result.section_options
    assert len(options) >= 2
    other = next(
        o
        for o in options
        if o.section.designation != ctx.fan_support_result.recommended_section.designation
    )
    chosen_ctx = context_for_section_choice(ctx, other.section.designation)
    assert any(w.code == "W-SEC-CHOICE" for w in chosen_ctx.warnings)
    assert chosen_ctx.fan_support_result.recommended_section.designation == (
        other.section.designation
    )
