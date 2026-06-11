"""Consistência entre PDF, Excel e CSV — o mesmo cálculo tem de contar a
mesma história nos três formatos (auditoria F2.4 / M-07)."""
import csv
import io

import openpyxl
import pytest
from conftest import pdf_text

from sfsc.engines.selector import run_full_calculation
from sfsc.enums import Country, FanConnectionType, FanType, SupportType
from sfsc.models import FanSupportInput, FanUnit
from sfsc.reports.exports import generate_csv, generate_excel
from sfsc.reports.memorial_pdf import generate_pdf


@pytest.fixture(scope="module")
def ctx():
    inp = FanSupportInput(
        project_name="Consistência", support_tag="FSU-CONS",
        fan_units=[FanUnit(fan_type=FanType.CENTRIFUGAL, weight_kg=280.0,
                           operating_weight_kg=300.0,
                           footprint_length_mm=1000.0, footprint_width_mm=800.0,
                           centre_of_gravity_height_mm=350.0)],
        support_type=SupportType.PEDESTAL,
        country=Country.PORTUGAL, seismic_zone="1.2",
        installation_height_mm=700.0, span_mm=1100.0,
        include_base_plate=True,
        fan_connection_type=FanConnectionType.DIRECT_FLANGE,
    )
    return run_full_calculation(inp)


@pytest.fixture(scope="module")
def outputs(ctx):
    csv_row = next(csv.DictReader(io.StringIO(generate_csv(ctx))))
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(ctx)))
    resumo = {
        str(row[0].value): row[1].value
        for row in wb["Resumo"].iter_rows(min_row=4)
        if row[0].value is not None
    }
    pdf = pdf_text(generate_pdf(ctx))
    return csv_row, resumo, pdf, wb


def test_status_consistent_across_outputs(ctx, outputs):
    csv_row, resumo, pdf, _ = outputs
    status = ctx.fan_support_result.status.value
    assert csv_row["status"] == status
    assert resumo["Estado global"] == status


def test_classification_consistent_across_outputs(ctx, outputs):
    csv_row, resumo, pdf, _ = outputs
    classification = ctx.fan_support_result.classification_level.value
    assert csv_row["classification"] == classification
    assert resumo["Classificação"] == classification
    assert classification.encode() in pdf


def test_section_consistent_across_outputs(ctx, outputs):
    csv_row, resumo, pdf, _ = outputs
    section = ctx.fan_support_result.recommended_section.designation
    assert csv_row["section"] == section
    assert resumo["Perfil seleccionado"] == section
    assert section.encode() in pdf


def test_governing_utilization_consistent(ctx, outputs):
    csv_row, resumo, _, _ = outputs
    eta = ctx.fan_support_result.section_verification.utilization_ratio
    assert float(csv_row["section_utilization"]) == pytest.approx(eta)
    assert float(resumo["Utilização máx. secção"]) == pytest.approx(eta)


def test_design_load_consistent(ctx, outputs):
    csv_row, resumo, _, _ = outputs
    load = ctx.fan_support_result.design_load_kN
    assert float(csv_row["design_load_kN"]) == pytest.approx(load)
    assert float(resumo["Carga de projecto [kN]"]) == pytest.approx(load)


def test_both_combination_levels_in_excel_and_pdf(ctx, outputs):
    """As tabelas de acções totais E de esforços no elemento aparecem nos dois
    formatos detalhados (correcção C-02)."""
    _, _, pdf, wb = outputs
    levels = {row[0].value for row in wb["Combinações"].iter_rows(min_row=2)
              if row[0].value}
    assert levels == {"Acções totais", "Esforços no elemento"}
    # Nos content streams o ReportLab escapa acentos em octal — usar
    # substrings ASCII dos títulos 4.1/4.2.
    assert b"4.1 Combina" in pdf
    assert b"4.2 Esfor" in pdf


def test_disclaimer_present_in_pdf(outputs):
    _, _, pdf, _ = outputs
    assert b"ENGINEERING ESTIMATE ONLY" in pdf
