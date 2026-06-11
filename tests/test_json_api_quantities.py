import io
import json

import openpyxl
import pytest

from sfsc.api import run_calculation
from sfsc.engines.quantities import calculate_quantities
from sfsc.engines.selector import run_full_calculation
from sfsc.enums import Country, FanConnectionType, FanType, SupportType
from sfsc.models import FanSupportInput, FanUnit
from sfsc.reports.export_json import export_report_dict, generate_json
from sfsc.reports.exports import generate_excel


def _inp(**updates):
    data = {
        "project_name": "JSON Contract",
        "support_tag": "FSU-JSON",
        "prepared_by": "Eng. JSON",
        "fan_units": [
            FanUnit(
                tag="FAN-1",
                fan_type=FanType.CENTRIFUGAL,
                weight_kg=250.0,
                operating_weight_kg=280.0,
                footprint_length_mm=1000.0,
                footprint_width_mm=800.0,
                centre_of_gravity_height_mm=350.0,
            )
        ],
        "support_type": SupportType.PEDESTAL,
        "country": Country.PORTUGAL,
        "seismic_zone": "1.2",
        "installation_height_mm": 700.0,
        "span_mm": 1000.0,
        "include_base_plate": True,
        "fan_connection_type": FanConnectionType.DIRECT_FLANGE,
    }
    data.update(updates)
    return FanSupportInput(**data)


def _payload():
    return _inp().model_dump(mode="json")


def test_json_export_has_required_contract_keys():
    ctx = run_full_calculation(_inp())
    payload = export_report_dict(ctx, calc_id="00000000-0000-4000-8000-000000000001")

    assert {
        "schema_version",
        "calc_id",
        "created_at",
        "project_id",
        "support_id",
        "software_version",
        "dataset_provenance",
        "input",
        "result",
        "assessment",
        "warnings",
        "citations",
        "assumptions",
        "limitations",
    } <= set(payload)
    assert payload["schema_version"] == "1.0"
    assert payload["support_id"] == "FSU-JSON"
    assert payload["result"]["quantities"]["estimate_note"].startswith("Estimativa")


def test_json_round_trip_preserves_core_values():
    ctx = run_full_calculation(_inp())
    parsed = json.loads(generate_json(ctx, calc_id="00000000-0000-4000-8000-000000000002"))
    res = ctx.fan_support_result

    assert parsed["input"]["support_tag"] == ctx.support_tag
    assert parsed["result"]["status"] == res.status.value
    assert (
        parsed["result"]["recommended_section"]["designation"]
        == res.recommended_section.designation
    )
    assert parsed["result"]["quantities"]["total_steel_mass_kg"] == pytest.approx(
        res.quantities.total_steel_mass_kg
    )


def test_quantities_manual_pedestal_formula():
    ctx = run_full_calculation(_inp(span_mm=1000.0))
    res = ctx.fan_support_result
    q = res.quantities

    assert q.structural_steel_length_m == pytest.approx(2.0)
    assert q.structural_steel_mass_kg == pytest.approx(2.0 * res.recommended_section.weight_kgm)
    assert q.plate_mass_kg > 0.0
    assert q.total_steel_mass_kg == pytest.approx(q.structural_steel_mass_kg + q.plate_mass_kg)
    assert q.anchor_count == res.anchor.n_anchors
    assert q.anchor_diameter_mm == res.anchor.anchor_diameter_mm
    assert q.weld_length_mm > 0.0


def test_quantities_can_be_recomputed_without_mutating_result():
    ctx = run_full_calculation(_inp())
    res = ctx.fan_support_result
    recomputed = calculate_quantities(ctx.fan_support_input, res)

    assert recomputed.model_dump() == res.quantities.model_dump()


def test_excel_contains_quantities_sheet():
    ctx = run_full_calculation(_inp())
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(ctx)))

    assert "Quantidades" in wb.sheetnames
    rows = {
        str(row[0].value): row[1].value
        for row in wb["Quantidades"].iter_rows(min_row=1)
        if row[0].value is not None
    }
    assert rows["Massa total aco [kg]"] == pytest.approx(
        ctx.fan_support_result.quantities.total_steel_mass_kg
    )


def test_api_success_returns_json_report_contract():
    response = run_calculation(_payload())

    assert response["ok"] is True
    report = response["report"]
    assert report["schema_version"] == "1.0"
    assert report["support_id"] == "FSU-JSON"
    assert report["result"]["quantities"]["anchor_count"] >= 1


def test_api_validation_error_is_structured():
    payload = _payload()
    del payload["project_name"]

    response = run_calculation(payload)

    assert response["ok"] is False
    assert response["errors"]
    assert {"field", "code", "message"} <= set(response["errors"][0])
    assert response["errors"][0]["field"] == "project_name"
